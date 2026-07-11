"""Генератор карты хронометража (Excel с живыми формулами). Версия v5 — согласована."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.properties import CalcProperties

GREEN="FF2E6E4E"; YELLOW="FFFFF2CC"; RESULT="FFDDF0E4"; SUB="FFEAF2ED"; ELEM="FF6E8B7B"; WHITE="FFFFFFFF"
thin=Side(style="thin", color="FFBFBFBF"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
KU_NORM={"machine":1.2,"machine_manual":1.5,"manual":2.0,"observation":2.5}
KU_LABEL={"machine":"машинные","machine_manual":"машинно-ручные","manual":"ручные","observation":"наблюдение"}
AOB_NOTE="из нормативов на обслуживание раб. места (межотраслевых/отраслевых) либо из вашего ФРД"
AOTL_NOTE="из нормативов на отдых и личные надобности (ОТЛ) либо из ФРД (доля отдыха в балансе дня)"

def bar(ws,row,text,fill=GREEN,size=11,span=5):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    c=ws.cell(row,1,text); c.font=Font(bold=True,color=WHITE,size=size)
    c.fill=PatternFill("solid",fgColor=fill); c.alignment=Alignment(vertical="center"); ws.row_dimensions[row].height=19

def kv(ws,row,label,value,note="",fill=None,bold=False,merge_val=False):
    b=ws.cell(row,2,label); b.font=Font(size=10,bold=bold)
    v=ws.cell(row,3,value); v.font=Font(size=10,bold=bold)
    if merge_val:
        ws.merge_cells(start_row=row,start_column=3,end_row=row,end_column=5)
        v.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    else:
        v.alignment=Alignment(horizontal="center")
    if fill:
        f=PatternFill("solid",fgColor=fill); b.fill=f; v.fill=f
    if note and not merge_val: ws.cell(row,4,note).font=Font(size=9,italic=True,color="FF7A7A7A")
    return v

def build(d):
    wt=d.get("work_type","manual"); kun=KU_NORM.get(wt,2.0)
    wb=openpyxl.Workbook()
    ws=wb.active; ws.title="Норма времени"
    dz=wb.create_sheet("Детализация"); mt=wb.create_sheet("Методика")
    ws.column_dimensions["A"].width=4; ws.column_dimensions["B"].width=44; ws.column_dimensions["C"].width=15; ws.column_dimensions["D"].width=30; ws.column_dimensions["E"].width=13
    dz.column_dimensions["A"].width=4; dz.column_dimensions["B"].width=44; dz.column_dimensions["C"].width=16; dz.column_dimensions["D"].width=18; dz.column_dimensions["E"].width=18
    for s in (ws,dz,mt): s.calculation=CalcProperties(fullCalcOnLoad=True); s.sheet_view.showGridLines=False

    # ── ДЕТАЛИЗАЦИЯ ──
    dz.cell(1,1,"ДЕТАЛИЗАЦИЯ ЗАМЕРОВ ПО ОПЕРАЦИЯМ"); dz.merge_cells("A1:E1")
    dz.cell(1,1).font=Font(bold=True,color=WHITE,size=12); dz.cell(1,1).fill=PatternFill("solid",fgColor=GREEN); dz.row_dimensions[1].height=24
    dz.cell(2,1,"Замеры по каждой операции. Отсюда формулы берут Тср для сводки на листе «Норма времени».")
    dz.cell(2,1).font=Font(size=9,italic=True,color="FF7A7A7A"); dz.merge_cells("A2:E2")
    r=4; refs=[]
    for ei,el in enumerate(d.get("elements",[]),1):
        bar(dz,r,f"Операция {ei}. {el.get('name','')}",fill=ELEM,size=10); r+=1
        dz.cell(r,1,"№"); dz.cell(r,2,"Длительность, мин"); dz.cell(r,3,"Учтён (1)/искл. (0)"); dz.cell(r,4,"Учтённая, мин")
        for c in (1,2,3,4):
            cc=dz.cell(r,c); cc.font=Font(bold=True,size=9); cc.fill=PatternFill("solid",fgColor=SUB); cc.alignment=Alignment(horizontal="center",wrap_text=True); cc.border=BORD
        first=r+1
        for j,m in enumerate(el.get("measurements",[]),1):
            r+=1
            dz.cell(r,1,j).alignment=Alignment(horizontal="center")
            dz.cell(r,2,m.get("min",0)).alignment=Alignment(horizontal="center")
            dz.cell(r,3,0 if m.get("excluded") else 1).alignment=Alignment(horizontal="center")
            dz.cell(r,4,f'=IF(C{r}=1,B{r},"")').alignment=Alignment(horizontal="center")
            for c in (1,2,3,4): dz.cell(r,c).border=BORD; dz.cell(r,c).font=Font(size=10)
        last=r; acc=f"D{first}:D{last}"
        r+=1; c_avg=r; kv(dz,r,"→ Тср операции (среднее по принятым), мин",f"=IF(COUNT({acc})=0,0,AVERAGE({acc}))",bold=True,fill=SUB)
        r+=1; c_ku=r; kv(dz,r,"→ Ку = Тмакс/Тмин",f'=IF(COUNT({acc})=0,"",MAX({acc})/MIN({acc}))')
        r+=1; kv(dz,r,"→ Устойчивость ряда",f'=IF(COUNT({acc})=0,"нет замеров",IF(MAX({acc})/MIN({acc})<={kun},"устойчив","неустойчив — исключите выбросы/добавьте замеры"))')
        r+=2; refs.append((el.get("name",""),c_avg,c_ku))

    # ── НОРМА ВРЕМЕНИ ──
    ws.cell(1,1,"НОРМА ВРЕМЕНИ · метод хронометража"); ws.merge_cells("A1:E1")
    ws.cell(1,1).font=Font(bold=True,color=WHITE,size=13); ws.cell(1,1).fill=PatternFill("solid",fgColor=GREEN); ws.row_dimensions[1].height=26
    ws.cell(3,2,"НОРМА ВРЕМЕНИ НА:"); ws.cell(3,2).font=Font(bold=True,size=11)
    ws.merge_cells("C3:E3"); nm=ws.cell(3,3,d.get("operation_name","")); nm.font=Font(bold=True,size=12); nm.fill=PatternFill("solid",fgColor=RESULT); ws.row_dimensions[3].height=22
    ws.cell(4,2,"вид работ / направление, на которое устанавливается норматив").font=Font(size=8.5,italic=True,color="FF7A7A7A")
    r=6; bar(ws,r,"ИСХОДНЫЕ ДАННЫЕ")
    r+=1; kv(ws,r,"Должность / профессия",d.get("position_name",""),merge_val=True)
    r+=1; kv(ws,r,"Подразделение",d.get("department_name",""),merge_val=True)
    r+=1; kv(ws,r,"Единица объёма работ",d.get("unit_of_work",""),merge_val=True)
    r+=1; kv(ws,r,"Тип работы",KU_LABEL.get(wt,wt),f"норматив устойчивости Ку ≤ {kun}")
    r+=1; kv(ws,r,"Наблюдатель / дата",f"{d.get('observer_name','')} · {d.get('obs_date','')}",merge_val=True)

    r+=2; bar(ws,r,"ИЗ ЧЕГО СОСТОИТ НОРМА")
    r+=1
    for c,t in zip((1,2,3,4),("№","Операция (составляющая)","Тср, мин","Ку")):
        cc=ws.cell(r,c,t); cc.font=Font(bold=True,size=10); cc.fill=PatternFill("solid",fgColor=SUB); cc.border=BORD; cc.alignment=Alignment(horizontal="center",wrap_text=True)
    sf=r+1
    for i,(name,ar,kr) in enumerate(refs,1):
        r+=1
        ws.cell(r,1,i).alignment=Alignment(horizontal="center"); ws.cell(r,1).border=BORD; ws.cell(r,1).font=Font(size=10)
        ws.cell(r,2,name).font=Font(size=10); ws.cell(r,2).border=BORD
        ws.cell(r,3,f"=Детализация!C{ar}").alignment=Alignment(horizontal="center"); ws.cell(r,3).border=BORD; ws.cell(r,3).font=Font(size=10)
        ws.cell(r,4,f"=Детализация!C{kr}").alignment=Alignment(horizontal="center"); ws.cell(r,4).border=BORD; ws.cell(r,4).font=Font(size=10)
    sl=r
    r+=1; c_top=r
    ws.cell(r,2,"Топ — оперативное время = сумма операций, мин").font=Font(bold=True,size=10.5)
    ws.cell(r,3,f"=SUM(C{sf}:C{sl})").font=Font(bold=True,size=10.5); ws.cell(r,3).alignment=Alignment(horizontal="center")
    for c in (1,2,3,4): ws.cell(r,c).fill=PatternFill("solid",fgColor=SUB); ws.cell(r,c).border=BORD

    r+=2; bar(ws,r,"РАСЧЁТ НОРМЫ")
    r+=1; c_aob=r; kv(ws,r,"аоб — обслуживание раб. места, %",d.get("servicing_percent",0),AOB_NOTE,fill=YELLOW)
    r+=1; c_aotl=r; kv(ws,r,"аотл — отдых и личные надобности, %",d.get("rest_percent",0),AOTL_NOTE,fill=YELLOW)
    r+=1; c_tsht=r; kv(ws,r,"Тшт = Топ × (1+(аоб+аотл)/100), мин",f"=C{c_top}*(1+(C{c_aob}+C{c_aotl})/100)",bold=True)
    r+=1; c_tpz=r; kv(ws,r,"Тпз — подготовит.-заключит. на партию, мин",d.get("prep_final_minutes",0),"из ваших нормативов; 0 — если не применяется",fill=YELLOW)
    r+=1; c_n=r; kv(ws,r,"n — размер партии",d.get("batch_size",""),"пусто — для непрерывных/единичных работ",fill=YELLOW)
    r+=1; c_tk=r; kv(ws,r,"Тшт-к = Тшт + Тпз/n, мин",f'=IF(OR(C{c_n}="",C{c_n}=0),C{c_tsht},C{c_tsht}+C{c_tpz}/C{c_n})',bold=True)
    r+=1
    ws.cell(r,2,"ИТОГОВАЯ НОРМА ВРЕМЕНИ, мин"); ws.cell(r,3,f"=C{c_tk}"); ws.cell(r,4,f"на 1 {d.get('unit_of_work','ед.')}")
    for c in (1,2,3,4): ws.cell(r,c).fill=PatternFill("solid",fgColor=RESULT); ws.cell(r,c).font=Font(bold=True,size=11.5)
    ws.cell(r,3).alignment=Alignment(horizontal="center"); ws.row_dimensions[r].height=20
    r+=1; kv(ws,r,"Источник надбавок",d.get("norms_source",""),merge_val=True)

    # ── МЕТОДИКА ──
    mt.column_dimensions["A"].width=3; mt.column_dimensions["B"].width=15; mt.column_dimensions["C"].width=70; mt.column_dimensions["D"].width=40
    mt.cell(1,1,"МЕТОДИКА · показатели, определения и формулы расчёта"); mt.merge_cells("A1:D1")
    mt.cell(1,1).font=Font(bold=True,color=WHITE,size=12); mt.cell(1,1).fill=PatternFill("solid",fgColor=GREEN); mt.row_dimensions[1].height=22
    mt.cell(2,2,"Норма времени устанавливается методом хронометража. Ниже — все величины, коэффициенты и формулы механизма расчёта.")
    mt.cell(2,2).font=Font(size=9,italic=True,color="FF7A7A7A"); mt.merge_cells("B2:D2")
    hr=4
    for c,t in zip((2,3,4),("Обозначение","Наименование и определение","Формула / источник")):
        cc=mt.cell(hr,c,t); cc.font=Font(bold=True,size=10,color=WHITE); cc.fill=PatternFill("solid",fgColor=ELEM); cc.border=BORD; cc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    rows=[
      ("Замер","Длительность выполнения операции при одном наблюдении (мин).","фиксируется нормировщиком"),
      ("Тмин, Тмакс","Наименьшая и наибольшая длительность среди ПРИНЯТЫХ замеров ряда.","из ряда замеров"),
      ("Тср","Среднее время операции-составляющей — среднее арифметическое принятых замеров.","Тср = среднее(принятые замеры)"),
      ("Ку","Коэффициент устойчивости хронометражного ряда — показывает степень колебания (разброса) замеров. Чем ближе к 1, тем стабильнее ряд. Ряд пригоден для нормирования, если Ку не превышает норматив.","Ку = Тмакс / Тмин"),
      ("Норматив Ку","Предельно допустимое значение Ку по типу работы. Если фактический Ку выше — ряд неустойчив, выброс исключают или добавляют замеры.","машинные ≤1.2 · машинно-ручные ≤1.5 · ручные ≤2.0 · наблюдение ≤2.5"),
      ("Топ","Оперативное время операции — время непосредственного её выполнения (основное + вспомогательное). В карте = сумма Тср всех операций-составляющих.","Топ = Σ Тср"),
      ("аоб","Надбавка на обслуживание рабочего места — время на уход за рабочим местом в течение смены, % от Топ.","из нормативов на обслуживание либо из ФРД"),
      ("аотл","Надбавка на отдых и личные надобности (ОТЛ) — регламентированное время на отдых, % от Топ.","из нормативов на ОТЛ либо из ФРД"),
      ("Тшт","Штучное время — норма времени на единицу работы с учётом надбавок на обслуживание и отдых.","Тшт = Топ × (1 + (аоб + аотл)/100)"),
      ("Тпз","Подготовительно-заключительное время — на подготовку к работе и её завершение; нормируется на всю партию, а не на единицу.","из нормативов"),
      ("n","Размер партии — число единиц, на которое распределяется Тпз.","задаётся пользователем"),
      ("Тшт-к","Штучно-калькуляционное время — штучное время плюс доля ПЗ, приходящаяся на единицу. Итоговая норма.","Тшт-к = Тшт + Тпз / n"),
    ]
    rr=hr
    for obn,name,form in rows:
        rr+=1
        mt.cell(rr,2,obn).font=Font(bold=True,size=10); mt.cell(rr,2).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); mt.cell(rr,2).border=BORD
        mt.cell(rr,3,name).font=Font(size=10); mt.cell(rr,3).alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); mt.cell(rr,3).border=BORD
        mt.cell(rr,4,form).font=Font(size=10); mt.cell(rr,4).alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); mt.cell(rr,4).border=BORD
    rr+=2
    mt.cell(rr,2,"Источник надбавок (аоб, аотл, Тпз):").font=Font(bold=True,size=10); mt.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=4)
    for txt in ["межотраслевые нормативы времени (утв. Минтруда / НИИ труда) — общие;",
                "отраслевые нормативы времени — по конкретной отрасли, если разработаны;",
                "собственный ФРД (фотография рабочего дня) — доля обслуживания и отдыха из реального баланса дня.",
                "Значения зависят от условий труда и НЕ выдумываются.","",
                "Итоговая норма (Тшт-к) оформляется локальным нормативным актом (ЛНА)."]:
        rr+=1; mt.cell(rr,2,("• "+txt) if txt else ""); mt.cell(rr,2).font=Font(size=10); mt.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=4); mt.cell(rr,2).alignment=Alignment(wrap_text=True,vertical="center")

    for s in (ws,dz):
        for rowc in s.iter_rows():
            for cell in rowc:
                al=cell.alignment
                cell.alignment=Alignment(horizontal=al.horizontal, vertical="center", wrap_text=True)
                if isinstance(cell.value,str) and cell.value.startswith("="): cell.number_format="0.00"
    return wb
