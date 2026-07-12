"""Генератор КЛИЕНТСКОГО Excel «Баланс рабочего времени + Lean-анализ» (ФРД). v2.
Без листа «Методика» (методика — внутренняя, клиенту не отдаётся). Только короткая легенда кодов."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.properties import CalcProperties

GREEN="FF2E6E4E"; YELLOW="FFFFF2CC"; RESULT="FFDDF0E4"; SUB="FFEAF2ED"; WHITE="FFFFFFFF"
RED="FFE24B4A"; AMBER="FFEF9F27"; BLUE="FF378ADD"
thin=Side(style="thin",color="FFBFBFBF"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)

TT_LABEL={"PV":"Продуктивное","UV":"Условно-продуктивное","LT":"Потери","PR":"Личное / отдых","EX":"Внешние факторы"}
TT_COLOR={"PV":"FF2E6E4E","UV":AMBER,"LT":RED,"PR":BLUE,"EX":"FF8A6D3B"}
VT_LABEL={"VA":"Создаёт ценность","BR":"Необходимо бизнесу","NVA":"Не создаёт ценность (потери)"}
VT_COLOR={"VA":"FF2E6E4E","BR":AMBER,"NVA":RED}
LST="Наблюдательный лист"

def bar(ws,row,text,fill=GREEN,size=11,span=6):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    c=ws.cell(row,1,text); c.font=Font(bold=True,color=WHITE,size=size)
    c.fill=PatternFill("solid",fgColor=fill); c.alignment=Alignment(vertical="center"); ws.row_dimensions[row].height=19

def kv(ws,row,label,value,fill=None,bold=False,note=""):
    b=ws.cell(row,2,label); b.font=Font(size=10,bold=bold)
    v=ws.cell(row,3,value); v.font=Font(size=10,bold=bold); v.alignment=Alignment(horizontal="center")
    if fill:
        f=PatternFill("solid",fgColor=fill); b.fill=f; v.fill=f
    if note: ws.cell(row,4,note).font=Font(size=9,italic=True,color="FF7A7A7A")
    return v

def build_frd(d):
    recs=d.get("records",[])
    wb=openpyxl.Workbook()
    ws=wb.active; ws.title="Баланс"
    lst=wb.create_sheet(LST)
    for s in (ws,lst): s.calculation=CalcProperties(fullCalcOnLoad=True); s.sheet_view.showGridLines=False
    ws.column_dimensions["A"].width=3
    for col,w in zip("BCDEF",[30,14,10,30,10]): ws.column_dimensions[col].width=w

    # ── НАБЛЮДАТЕЛЬНЫЙ ЛИСТ ──
    lst.column_dimensions["A"].width=4
    for col,w in zip("BCDEFG",[11,11,11,46,18,20]): lst.column_dimensions[col].width=w
    lst.cell(1,1,"НАБЛЮДАТЕЛЬНЫЙ ЛИСТ ФРД"); lst.merge_cells("A1:G1")
    lst.cell(1,1).font=Font(bold=True,color=WHITE,size=12); lst.cell(1,1).fill=PatternFill("solid",fgColor=GREEN); lst.row_dimensions[1].height=22
    for c,t in enumerate(["№","Начало","Оконч.","Длит.,мин","Активность","Вид времени","Ценность"],1):
        cc=lst.cell(3,c,t); cc.font=Font(bold=True,size=10); cc.fill=PatternFill("solid",fgColor=SUB); cc.border=BORD; cc.alignment=Alignment(horizontal="center",wrap_text=True,vertical="center")
    r=4
    for i,rec in enumerate(recs,1):
        tt=str(rec.get("time_type","")).strip(); vt=str(rec.get("value_type","")).strip()
        lst.cell(r,1,i).alignment=Alignment(horizontal="center")
        lst.cell(r,2,rec.get("start","")).alignment=Alignment(horizontal="center")
        lst.cell(r,3,rec.get("end","")).alignment=Alignment(horizontal="center")
        lst.cell(r,4,round(float(rec.get("duration",0) or 0),1)).alignment=Alignment(horizontal="center")
        lst.cell(r,5,rec.get("activity",""))
        e=lst.cell(r,6,f"{tt} — {TT_LABEL.get(tt,tt)}"); e.font=Font(size=9,color=TT_COLOR.get(tt,"FF000000"),bold=True)
        g=lst.cell(r,7,f"{vt} — {VT_LABEL.get(vt,vt)}"); g.font=Font(size=9,color=VT_COLOR.get(vt,"FF000000"),bold=True)
        lst.cell(r,8,tt); lst.cell(r,9,vt)
        for c in range(1,6): lst.cell(r,c).font=Font(size=10)
        for c in range(1,8): lst.cell(r,c).border=BORD
        r+=1
    last=r-1; first=4
    lst.column_dimensions["H"].hidden=True; lst.column_dimensions["I"].hidden=True
    dur=f"'{LST}'!$D${first}:$D${last}"; ttc=f"'{LST}'!$H${first}:$H${last}"; vtc=f"'{LST}'!$I${first}:$I${last}"

    # ── БАЛАНС ──
    ws.cell(1,1,"БАЛАНС РАБОЧЕГО ВРЕМЕНИ · ФРД + Lean-анализ"); ws.merge_cells("A1:F1")
    ws.cell(1,1).font=Font(bold=True,color=WHITE,size=13); ws.cell(1,1).fill=PatternFill("solid",fgColor=GREEN); ws.row_dimensions[1].height=26
    r=3; bar(ws,r,"ИСХОДНЫЕ ДАННЫЕ")
    r+=1; kv(ws,r,"Должность",d.get("position",""))
    r+=1; kv(ws,r,"Подразделение",d.get("department",""))
    r+=1; kv(ws,r,"Дата наблюдения",d.get("obs_date",""))
    r+=1; kv(ws,r,"Смена",f"{d.get('work_start','')} – {d.get('work_end','')}")
    r+=1; tot_row=r; kv(ws,r,"Наблюдаемое время, мин",f"=SUM({dur})",bold=True,fill=SUB)
    TOT=f"$C${tot_row}"

    r+=2; bar(ws,r,"БАЛАНС ПО ВИДАМ ВРЕМЕНИ")
    r+=1
    for c,t in zip((2,3,4),("Вид времени","Мин","% дня")):
        cc=ws.cell(r,c,t); cc.font=Font(bold=True,size=10); cc.fill=PatternFill("solid",fgColor=SUB); cc.border=BORD; cc.alignment=Alignment(horizontal="center")
    tt_rows={}
    for code in ("PV","UV","LT","PR","EX"):
        r+=1; tt_rows[code]=r
        ws.cell(r,2,f"{code} · {TT_LABEL[code]}").font=Font(size=10,color=TT_COLOR[code],bold=True)
        ws.cell(r,3,f'=SUMIF({ttc},"{code}",{dur})').alignment=Alignment(horizontal="center")
        ws.cell(r,4,f'=IF({TOT}=0,0,C{r}/{TOT})').alignment=Alignment(horizontal="center"); ws.cell(r,4).number_format="0.0%"
        for c in (2,3,4): ws.cell(r,c).border=BORD

    r+=2; bar(ws,r,"КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ")
    pv=f"C{tt_rows['PV']}"; uv=f"C{tt_rows['UV']}"; lt=f"C{tt_rows['LT']}"; pr=f"C{tt_rows['PR']}"; ex=f"C{tt_rows['EX']}"
    losses=f"({lt}+{ex})"
    r+=1; kv(ws,r,"Кисп — коэффициент использования",f'=IF({TOT}=0,0,({pv}+{uv}+{pr})/{TOT})',bold=True,fill=RESULT); ws.cell(r,3).number_format="0.0%"
    r+=1; kv(ws,r,"Доля потерь (LT+EX)",f'=IF({TOT}=0,0,{losses}/{TOT})',fill=YELLOW); ws.cell(r,3).number_format="0.0%"
    r+=1; kv(ws,r,"ППТ — возможный рост производительности",f'=IF(({pv}+{uv})=0,0,{losses}/({pv}+{uv}))',bold=True,fill=RESULT,note="за счёт устранения потерь"); ws.cell(r,3).number_format="0.0%"

    r+=2; bar(ws,r,"LEAN-СРЕЗ ПО ЦЕННОСТИ")
    r+=1
    for c,t in zip((2,3,4),("Категория ценности","Мин","% дня")):
        cc=ws.cell(r,c,t); cc.font=Font(bold=True,size=10); cc.fill=PatternFill("solid",fgColor=SUB); cc.border=BORD; cc.alignment=Alignment(horizontal="center")
    for code in ("VA","BR","NVA"):
        r+=1
        ws.cell(r,2,f"{code} · {VT_LABEL[code]}").font=Font(size=10,color=VT_COLOR[code],bold=True)
        ws.cell(r,3,f'=SUMIF({vtc},"{code}",{dur})').alignment=Alignment(horizontal="center")
        ws.cell(r,4,f'=IF({TOT}=0,0,C{r}/{TOT})').alignment=Alignment(horizontal="center"); ws.cell(r,4).number_format="0.0%"
        for c in (2,3,4): ws.cell(r,c).border=BORD

    r+=2; bar(ws,r,"ПРОЕКТИРУЕМЫЙ (НОРМАЛЬНЫЙ) БАЛАНС")
    r+=1; kv(ws,r,"Время после устранения потерь, мин",f'={TOT}-{losses}',fill=SUB)
    r+=1; kv(ws,r,"Высвобождаемый резерв, мин",f'={losses}',bold=True,fill=RESULT,note="потенциал для полезной работы")

    # ── КОРОТКАЯ ЛЕГЕНДА (только расшифровка кодов, без методики) ──
    r+=2; bar(ws,r,"ЛЕГЕНДА")
    r+=1
    ws.cell(r,2,"Вид времени: PV — продуктивное · UV — условно-продуктивное · LT — потери · PR — личное/отдых · EX — внешние факторы")
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
    ws.cell(r,2).font=Font(size=9,color="FF5A5A5A"); ws.cell(r,2).alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[r].height=28
    r+=1
    ws.cell(r,2,"Ценность (Lean): VA — создаёт ценность · BR — необходимо бизнесу · NVA — не создаёт ценности (потери)")
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
    ws.cell(r,2).font=Font(size=9,color="FF5A5A5A"); ws.cell(r,2).alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[r].height=28

    for s in (ws,lst):
        for rowc in s.iter_rows():
            for cell in rowc:
                al=cell.alignment; cell.alignment=Alignment(horizontal=al.horizontal,vertical="center",wrap_text=True)
                if isinstance(cell.value,str) and cell.value.startswith("=") and cell.number_format=="General": cell.number_format="0.0"
    return wb

if __name__=="__main__":
    import io
    sample={"position":"Специалист по кадрам","department":"Отдел кадров","obs_date":"12.03.2026","work_start":"08:30","work_end":"17:30",
      "records":[
        {"start":"08:30","end":"08:45","duration":15,"activity":"Настройка рабочего места, запуск 1С","time_type":"PR","value_type":"BR"},
        {"start":"08:45","end":"09:30","duration":45,"activity":"Оформление приказов о приёме","time_type":"PV","value_type":"VA"},
        {"start":"09:30","end":"10:10","duration":40,"activity":"Поиск кадрового документа в архиве","time_type":"LT","value_type":"NVA"},
        {"start":"10:10","end":"11:00","duration":50,"activity":"Участие в совещании у руководителя","time_type":"UV","value_type":"BR"},
        {"start":"11:00","end":"11:15","duration":15,"activity":"Перерыв","time_type":"PR","value_type":"NVA"},
        {"start":"11:15","end":"12:00","duration":45,"activity":"Ожидание согласования заявки","time_type":"EX","value_type":"NVA"},
      ]}
    wb=build_frd(sample); buf=io.BytesIO(); wb.save(buf)
    print("OK, размер:",len(buf.getvalue()),"| листов:",len(wb.sheetnames),"|",wb.sheetnames)
    wb.save("/mnt/user-data/outputs/FRD_BALANCE_sample.xlsx")
